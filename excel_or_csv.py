import openpyxl
import csv


def list_to_excel(to_excel_list: list , output_excel_path: str = "output.xlsx"):
    """ 多次元リストをExcelファイルに転記する関数（新転上書き保存） """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # 多次元リストのサイズを取得(行ごとで列数に違いがあることを考慮)
    row_num , col_num = len(to_excel_list) , 0
    for row in range(row_num):
        predict_col = len(to_excel_list[row])
        if predict_col > col_num:
            col_num = predict_col

    for row in range(row_num):
        for col in range(col_num):
            try:
                sheet.cell(row=row+1, column=col+1).value = to_excel_list[row][col]
            except IndexError:
                pass
    workbook.save(output_excel_path)



def many_list_to_excel(many_list: list , output_excel_path: str = "output.xlsx"):
    """ 多次元リストを複数含むリストをExcelファイルの複数のシートにそれぞれ転記する関数 """
    workbook = openpyxl.Workbook()

    for sheet_index, to_excel_list in enumerate(many_list):
        sheet_name = f"Sheet{sheet_index + 1}"
        sheet = workbook.create_sheet(title=sheet_name)

        # 多次元リストのサイズを取得(行ごとで列数に違いがあることを考慮)
        row_num, col_num = len(to_excel_list), 0
        for row in range(row_num):
            predict_col = len(to_excel_list[row])
            if predict_col > col_num:
                col_num = predict_col

        for row in range(row_num):
            for col in range(col_num):
                try:
                    sheet.cell(row=row+1, column=col+1).value = to_excel_list[row][col]
                except IndexError:
                    pass

    # 最初のデフォルトのシートを削除
    default_sheet = workbook.get_sheet_by_name("Sheet")
    workbook.remove_sheet(default_sheet)

    workbook.save(output_excel_path)


def excel_to_order_number_list(input_excel_path: str = "output.xlsx"):
    """ 多次元リストをExcelファイルに転記する関数（新転上書き保存） """
    workbook = openpyxl.load_workbook(input_excel_path)
    sheet = workbook.active
    past_order_number_list = []
    past_participant_number_list = []
    for row in sheet.iter_rows(min_row = 1, max_col = 9, values_only = True):
        order_number = row[8]
        if order_number is not None:
            past_order_number_list.append(order_number)
    return past_order_number_list




def csv_to_list(csv_path: str = "output.csv"):
    """ 多次元データを含むcsvからリストに変換する関数
            改行ありの文字列を含む場合はバグる
    """
    data_list = []
    with open(csv_path, 'r' , encoding="utf-8-sig") as file:
        csv_reader = csv.reader(file)
        for row in csv_reader:
            data_list.append(row)
    return data_list


def list_to_csv(to_csv_list: list , csv_path: str = "output.csv"):
    """ 多次元リストのデータをcsvファイルに保存する関数
            改行ありの文字列を含む場合はバグる
    """
    with open(csv_path, 'w' , encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(to_csv_list)


def many_list_to_csv(many_list: list, base_csv_path: str = "output"):
    """ 多次元リストを複数含むリストを複数のcsvに転記する関数 """
    csv_path_list = []
    for index, to_csv_list in enumerate(many_list):
        csv_path = f"{base_csv_path}_{index + 1}.csv"
        csv_path_list.append(csv_path)
        with open(csv_path, 'w', newline='', encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerows(to_csv_list)
    return csv_path_list
